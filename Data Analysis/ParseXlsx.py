import openpyxl
import json
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("C:\\Users\\Yoni\\RiderProjects\\Hackathon-Project-2022\\Hackathon-Project-2022\\Data Analysis\\diseases.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

# Iterate the loop to read the cell values
d = dict()
for col in dataframe1.iter_cols(1, dataframe1.max_column):
    d[col[0].value] = []
    for row in range(1, dataframe1.max_row):
        if col[row].value:
            d[col[0].value].append(col[row].value)

json_object = json.dumps(d)
s = json_object.__str__()
print(type(s))

# Serializing json



with open("C:\\Users\\Yoni\\RiderProjects\\Hackathon-Project-2022\\Hackathon-Project-2022\\Data Analysis\\xl.txt", "w") as f:
    f.write(s)
